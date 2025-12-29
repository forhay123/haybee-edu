# python-service/app/domains/individual_processing/scheme_extractor.py
from typing import List
from .schemas import SchemeTopic
import pdfplumber
from openpyxl import load_workbook
from app.core.logger import get_logger

logger = get_logger(__name__)

def extract_scheme(file_url: str) -> List[SchemeTopic]:
    """
    Extracts topics, subtopics, learning objectives from PDF or Excel.
    """
    logger.info(f"📄 Extracting scheme from: {file_url}")
    
    if file_url.endswith(".pdf"):
        return _extract_scheme_pdf(file_url)
    elif file_url.endswith(".xlsx") or file_url.endswith(".xls"):
        return _extract_scheme_excel(file_url)
    else:
        raise ValueError(f"Unsupported file type: {file_url}")


def _extract_scheme_pdf(file_url: str) -> List[SchemeTopic]:
    topics = []
    with pdfplumber.open(file_url) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            for line in text.split("\n"):
                if line.strip():
                    # naive: "Week Topic | Subject | Subtopics | Objectives"
                    parts = line.split("|")
                    if len(parts) >= 2:
                        week_topic = parts[0].split(" ", 1)
                        week = int(week_topic[0])
                        topic = week_topic[1] if len(week_topic) > 1 else "Unknown"
                        subject = parts[1].strip() if len(parts) > 1 else "Unknown Subject"
                        subtopics = parts[2].split(",") if len(parts) > 2 else []
                        objectives = parts[3].split(",") if len(parts) > 3 else None
                        
                        topics.append(SchemeTopic(
                            week_number=week,  # ✅ Use week_number (not week)
                            topic=topic,
                            subject=subject,  # ✅ Required field
                            subtopics=subtopics,
                            objectives=objectives  # ✅ Use objectives (not learning_objectives)
                        ))
    
    logger.info(f"✅ Extracted {len(topics)} topics from PDF")
    return topics


def _extract_scheme_excel(file_url: str) -> List[SchemeTopic]:
    topics = []
    wb = load_workbook(filename=file_url)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Expected columns: Week | Topic | Subject | Subtopics | Objectives
            if len(row) >= 3:
                week = int(row[0])
                topic = str(row[1])
                subject = str(row[2]) if row[2] else "Unknown Subject"
                subtopics = row[3].split(",") if len(row) > 3 and row[3] else []
                objectives = row[4].split(",") if len(row) > 4 and row[4] else None
                
                topics.append(SchemeTopic(
                    week_number=week,  # ✅ Use week_number
                    topic=topic,
                    subject=subject,  # ✅ Required field
                    subtopics=subtopics,
                    objectives=objectives  # ✅ Use objectives
                ))
    
    wb.close()
    logger.info(f"✅ Extracted {len(topics)} topics from Excel")
    return topics