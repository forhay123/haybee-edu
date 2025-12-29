"""
app/domains/individual_processing/timetable_extractor.py
✅ SPRINT 9: Enhanced with conflict detection and validation
"""
from typing import List, Set, Dict, Tuple, Optional
import os
import tempfile
import requests
from pathlib import Path
import re
import cv2
import numpy as np
import pdfplumber
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from datetime import time, timedelta

from app.core.logger import get_logger
from app.domains.individual_processing.schemas import TimetableEntry
import signal
from contextlib import contextmanager

logger = get_logger(__name__)

class TimeoutException(Exception):
    pass

@contextmanager
def time_limit(seconds):
    """Context manager for timeouts"""
    def signal_handler(signum, frame):
        raise TimeoutException("Timed out!")
    
    signal.signal(signal.SIGALRM, signal_handler)
    signal.alarm(seconds)
    try:
        yield
    finally:
        signal.alarm(0)

# Days of the week
DAYS_OF_WEEK = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']

# Known subjects (enhanced) - Academic subjects only
KNOWN_SUBJECTS = {
    # Core subjects
    'Mathematics', 'English', 'English Language', 'Physics', 'Chemistry',
    'Biology', 'Geography', 'History', 'Economics', 'Government',
    
    # Commercial subjects
    'ICT', 'Computer Studies', 'Accounting', 'Financial Accounting',
    'Office Practice', 'Book Keeping', 'Commerce', 'Bookkeeping',
    'ICT Practical',
    
    # Languages & Arts
    'Literature', 'Literature in English', 'French', 'Arabic', 'Yoruba',
    'Music', 'Art', 'Visual Arts', 'Performing Arts',
    
    # Religious/Civic
    'Civic Education', 'CRS', 'IRS', 'Christian Religious Studies',
    'Islamic Religious Studies', 'Civic', 'CRS/IRS', 'CRS / IRS',
    
    # Physical/Life
    'PHE', 'Physical Education', 'Physical and Health Education',
    'Agriculture', 'Home Economics', 'Agricultural Science',
    
    # Social Studies
    'Social Studies', 'Human and Social Biology', 'Health Science',
    
    # Technical
    'Technical Drawing', 'Basic Technology', 'Metalwork', 'Woodwork',
    'Auto Mechanics', 'Electrical Installation',
    
    # Others
    'Entrepreneurship', 'Guidance Counselling', 'Cultural Arts',
    'Information Technology', 'Business Studies'
}

# Non-academic periods to skip (not real subjects)
NON_ACADEMIC_PERIODS = {
    'Break', 'Closing', 'Free Period', 'Club Activities', 'Assembly',
    'Lunch', 'Recess', 'Morning Assembly', 'Devotion', 'Dismissal'
}

# Standard period times
PERIOD_TIMES = {
    1: ("08:00", "08:40"),
    2: ("08:40", "09:20"),
    3: ("09:20", "10:00"),
    4: ("10:20", "11:00"),
    5: ("11:00", "11:40"),
    6: ("11:40", "12:20"),
    7: ("13:00", "13:40"),
    8: ("13:40", "14:20"),
}


# ============================================================
# ✅ NEW: CONFLICT DETECTION CLASSES
# ============================================================

class ConflictType:
    """Conflict type constants"""
    TIME_OVERLAP = "TIME_OVERLAP"
    DUPLICATE_SUBJECT = "DUPLICATE_SUBJECT"
    INVALID_TIME_RANGE = "INVALID_TIME_RANGE"
    UNREALISTIC_DURATION = "UNREALISTIC_DURATION"
    TOO_MANY_PERIODS = "TOO_MANY_PERIODS"


class Conflict:
    """Represents a schedule conflict"""
    def __init__(self, conflict_type: str, day: str, severity: str, description: str,
                 entry1: Optional[Dict] = None, entry2: Optional[Dict] = None):
        self.type = conflict_type
        self.day = day
        self.severity = severity
        self.description = description
        self.entry1 = entry1
        self.entry2 = entry2
    
    def to_dict(self) -> Dict:
        """Convert to dictionary"""
        return {
            "type": self.type,
            "day": self.day,
            "severity": self.severity,
            "description": self.description,
            "entry1": self.entry1,
            "entry2": self.entry2
        }


def detect_conflicts(entries: List[TimetableEntry]) -> List[Conflict]:
    """
    ✅ NEW: Detect all types of conflicts in extracted timetable
    """
    logger.info("🔍 Detecting conflicts in timetable...")
    
    conflicts = []
    
    # 1. Detect time overlaps
    conflicts.extend(_detect_time_overlaps(entries))
    
    # 2. Detect duplicate subjects
    conflicts.extend(_detect_duplicate_subjects(entries))
    
    # 3. Detect invalid time ranges
    conflicts.extend(_detect_invalid_time_ranges(entries))
    
    # 4. Detect unrealistic schedules
    conflicts.extend(_detect_unrealistic_schedules(entries))
    
    logger.info(f"✅ Found {len(conflicts)} conflicts")
    
    # Log conflicts by type
    by_type = {}
    for conflict in conflicts:
        by_type[conflict.type] = by_type.get(conflict.type, 0) + 1
    
    for ctype, count in by_type.items():
        logger.info(f"   {ctype}: {count}")
    
    return conflicts


def _detect_time_overlaps(entries: List[TimetableEntry]) -> List[Conflict]:
    """Detect time overlap conflicts"""
    conflicts = []
    
    # Group by day
    by_day = {}
    for entry in entries:
        day = entry.day.upper()
        if day not in by_day:
            by_day[day] = []
        by_day[day].append(entry)
    
    # Check each day for overlaps
    for day, day_entries in by_day.items():
        # Sort by start time
        sorted_entries = sorted(day_entries, key=lambda e: e.start_time)
        
        for i in range(len(sorted_entries) - 1):
            for j in range(i + 1, len(sorted_entries)):
                entry1 = sorted_entries[i]
                entry2 = sorted_entries[j]
                
                if _has_time_overlap(entry1, entry2):
                    conflict = Conflict(
                        conflict_type=ConflictType.TIME_OVERLAP,
                        day=day,
                        severity="HIGH",
                        description=f"{entry1.subject} ({entry1.start_time}-{entry1.end_time}) "
                                  f"overlaps with {entry2.subject} ({entry2.start_time}-{entry2.end_time})",
                        entry1=_entry_to_dict(entry1),
                        entry2=_entry_to_dict(entry2)
                    )
                    conflicts.append(conflict)
                    logger.warning(f"⚠️ {conflict.description}")
    
    return conflicts


def _detect_duplicate_subjects(entries: List[TimetableEntry]) -> List[Conflict]:
    """Detect duplicate subjects on same day"""
    conflicts = []
    
    # Group by day
    by_day = {}
    for entry in entries:
        day = entry.day.upper()
        if day not in by_day:
            by_day[day] = []
        by_day[day].append(entry)
    
    # Check each day for duplicates
    for day, day_entries in by_day.items():
        # Group by subject
        by_subject = {}
        for entry in day_entries:
            subj = entry.subject
            if subj not in by_subject:
                by_subject[subj] = []
            by_subject[subj].append(entry)
        
        # Check for subjects appearing more than twice (more than 2 is suspicious)
        for subject, subj_entries in by_subject.items():
            if len(subj_entries) > 2:
                conflict = Conflict(
                    conflict_type=ConflictType.DUPLICATE_SUBJECT,
                    day=day,
                    severity="MEDIUM",
                    description=f"{subject} appears {len(subj_entries)} times on {day} - "
                              f"may be intentional for practical sessions",
                    entry1=_entry_to_dict(subj_entries[0]),
                    entry2=_entry_to_dict(subj_entries[1])
                )
                conflicts.append(conflict)
                logger.info(f"ℹ️ {conflict.description}")
    
    return conflicts


def _detect_invalid_time_ranges(entries: List[TimetableEntry]) -> List[Conflict]:
    """Detect invalid time ranges (end before start, unrealistic durations)"""
    conflicts = []
    
    for entry in entries:
        start = _parse_time(entry.start_time)
        end = _parse_time(entry.end_time)
        
        if not start or not end:
            continue
        
        # Check if end time is before or equal to start time
        if end <= start:
            conflict = Conflict(
                conflict_type=ConflictType.INVALID_TIME_RANGE,
                day=entry.day.upper(),
                severity="HIGH",
                description=f"Invalid time range for {entry.subject}: end time ({entry.end_time}) "
                          f"is before or equal to start time ({entry.start_time})",
                entry1=_entry_to_dict(entry)
            )
            conflicts.append(conflict)
            logger.warning(f"⚠️ {conflict.description}")
        
        # Check for unrealistic duration (> 2 hours for single period)
        duration_minutes = _time_diff_minutes(start, end)
        if duration_minutes > 120:
            conflict = Conflict(
                conflict_type=ConflictType.UNREALISTIC_DURATION,
                day=entry.day.upper(),
                severity="MEDIUM",
                description=f"Unusually long period for {entry.subject}: {duration_minutes} minutes "
                          f"({entry.start_time} - {entry.end_time}) - verify this is correct",
                entry1=_entry_to_dict(entry)
            )
            conflicts.append(conflict)
            logger.info(f"ℹ️ {conflict.description}")
    
    return conflicts


def _detect_unrealistic_schedules(entries: List[TimetableEntry]) -> List[Conflict]:
    """Detect unrealistic schedules (too many periods per day)"""
    conflicts = []
    
    # Group by day
    by_day = {}
    for entry in entries:
        day = entry.day.upper()
        if day not in by_day:
            by_day[day] = []
        by_day[day].append(entry)
    
    # Check for days with too many periods
    for day, day_entries in by_day.items():
        if len(day_entries) > 10:  # Most schools have max 8-9 periods
            conflict = Conflict(
                conflict_type=ConflictType.TOO_MANY_PERIODS,
                day=day,
                severity="MEDIUM",
                description=f"{day} has {len(day_entries)} periods scheduled - "
                          f"this seems unusually high. Please verify."
            )
            conflicts.append(conflict)
            logger.warning(f"⚠️ {conflict.description}")
    
    return conflicts


def _has_time_overlap(entry1: TimetableEntry, entry2: TimetableEntry) -> bool:
    """Check if two entries have overlapping time slots"""
    start1 = _parse_time(entry1.start_time)
    end1 = _parse_time(entry1.end_time)
    start2 = _parse_time(entry2.start_time)
    end2 = _parse_time(entry2.end_time)
    
    if not all([start1, end1, start2, end2]):
        return False
    
    # Two intervals overlap if: start1 < end2 AND start2 < end1
    return start1 < end2 and start2 < end1


def _parse_time(time_str: str) -> Optional[time]:
    """Parse time string to time object"""
    if not time_str:
        return None
    
    try:
        # Handle formats: "08:00", "8:00", "08:00:00"
        parts = time_str.split(':')
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        return time(hour, minute)
    except Exception as e:
        logger.warning(f"⚠️ Failed to parse time: {time_str} - {e}")
        return None


def _time_diff_minutes(start: time, end: time) -> int:
    """Calculate time difference in minutes"""
    start_minutes = start.hour * 60 + start.minute
    end_minutes = end.hour * 60 + end.minute
    return end_minutes - start_minutes


def _entry_to_dict(entry: TimetableEntry) -> Dict:
    """Convert TimetableEntry to dictionary"""
    return {
        "day": entry.day,
        "period_number": entry.period_number,
        "start_time": entry.start_time,
        "end_time": entry.end_time,
        "subject": entry.subject,
        "level": entry.level
    }


def validate_timetable(entries: List[TimetableEntry]) -> Tuple[bool, List[Conflict]]:
    """
    ✅ NEW: Validate complete timetable
    Returns: (is_valid, conflicts)
    """
    logger.info("✅ Validating timetable...")
    
    conflicts = detect_conflicts(entries)
    
    # Count high severity conflicts
    high_severity = [c for c in conflicts if c.severity == "HIGH"]
    
    is_valid = len(high_severity) == 0
    
    if is_valid:
        logger.info("✅ Timetable is valid (no high-severity conflicts)")
    else:
        logger.warning(f"⚠️ Timetable has {len(high_severity)} high-severity conflicts")
    
    return is_valid, conflicts


# ============================================================
# EXISTING EXTRACTION FUNCTIONS (PRESERVED)
# ============================================================

def extract_timetable(file_url: str) -> List[TimetableEntry]:
    """
    Main entry point for timetable extraction.
    Extracts COMPLETE timetable structure with days, periods, times, and subjects.
    ✅ ENHANCED: Now includes conflict detection
    """
    logger.info(f"🚀 Starting timetable extraction from: {file_url}")
    
    file_path, is_temp = _get_file_path(file_url)
    
    try:
        if file_url.lower().endswith('.pdf'):
            entries = _extract_from_pdf(file_path)
        elif file_url.lower().endswith(('.xlsx', '.xls')):
            entries = _extract_from_excel(file_path)
        elif file_url.lower().endswith(('.png', '.jpg', '.jpeg')):
            entries = _extract_from_image(file_path)
        elif file_url.lower().endswith(('.doc', '.docx', '.txt')):
            entries = _extract_from_text_document(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_url}")
        
        logger.info(f"📊 EXTRACTION SUMMARY:")
        logger.info(f"   Total entries extracted: {len(entries)}")
        
        # Group by day for logging
        by_day = {}
        for entry in entries:
            day = entry.day
            if day not in by_day:
                by_day[day] = []
            by_day[day].append(entry.subject)
        
        for day in DAYS_OF_WEEK:
            if day in by_day:
                logger.info(f"   {day}: {len(by_day[day])} periods - {', '.join(by_day[day][:3])}{'...' if len(by_day[day]) > 3 else ''}")
        
        # ✅ NEW: Detect conflicts
        is_valid, conflicts = validate_timetable(entries)
        
        if conflicts:
            logger.warning(f"⚠️ Detected {len(conflicts)} conflicts in extracted timetable")
            for conflict in conflicts[:5]:  # Log first 5 conflicts
                logger.warning(f"   {conflict.description}")
            if len(conflicts) > 5:
                logger.warning(f"   ... and {len(conflicts) - 5} more conflicts")
        
        return entries
        
    finally:
        if is_temp and os.path.exists(file_path):
            os.remove(file_path)


def _get_file_path(file_url: str) -> tuple:
    """Get file path from URL or local path."""
    if file_url.startswith(('http://', 'https://')):
        logger.info(f"⬇️ Downloading file from URL: {file_url}")
        
        response = requests.get(file_url, timeout=30)
        response.raise_for_status()
        
        ext = Path(file_url).suffix or '.pdf'
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        temp_file.write(response.content)
        temp_file.close()
        
        logger.info(f"✅ Downloaded to: {temp_file.name}")
        return temp_file.name, True
    else:
        logger.info(f"📁 Using local file: {file_url}")
        
        if not os.path.exists(file_url):
            raise FileNotFoundError(f"File not found: {file_url}")
        
        return file_url, False


def _extract_from_pdf(file_path: str) -> List[TimetableEntry]:
    """Extract complete timetable structure from PDF."""
    logger.info("📖 Starting PDF parsing...")
    
    with pdfplumber.open(file_path) as pdf:
        all_text = ""
        
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                all_text += text + "\n"
                logger.info(f"📄 Page {page_num + 1}: {len(text)} characters")
    
    # CRITICAL: Log the extracted text to see what we're working with
    logger.info(f"📝 Total extracted text length: {len(all_text)} characters")
    logger.debug(f"📝 First 500 chars: {all_text[:500]}")
    
    # Try structured parsing first
    entries = _parse_structured_timetable(all_text)
    
    if entries:
        logger.info(f"✅ Parsed {len(entries)} entries from structured format")
        return entries
    
    logger.warning("⚠️ Structured parsing failed, falling back to basic extraction")
    return _fallback_extraction(all_text)


def _parse_structured_timetable(text: str) -> List[TimetableEntry]:
    """
    Parse structured timetable text format.
    Now handles multiple period formats and extracts ALL periods.
    """
    entries = []
    current_day = None
    
    # Clean text
    text = _clean_ocr_text(text)
    lines = text.split('\n')
    
    logger.info(f"📝 Parsing {len(lines)} lines...")
    
    # ADD THIS LINE TO SEE THE ACTUAL TEXT:
    logger.info(f"📝 FULL EXTRACTED TEXT:\n{text}")
    
    for line_num, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
        
        # Check if this line is a day header
        line_upper = line.upper()
        for day in DAYS_OF_WEEK:
            if day in line_upper and len(line) < 30:  # Day headers are usually short
                current_day = day
                logger.info(f"📅 Found day: {day} at line {line_num}")
                break
        
        # Check if this line contains a period and subject
        if current_day:
            # Pattern 1: "1st Period: English Language"
            # Pattern 2: "1st Period — 8:00–8:40"
            # Pattern 3: "Period 1: English"
            period_patterns = [
                r'(\d+)(?:st|nd|rd|th)\s+Period[\s:—\-]+(.+)',  # "1st Period: Subject"
                r'Period\s+(\d+)[\s:—\-]+(.+)',  # "Period 1: Subject"
            ]
            
            matched = False
            for pattern in period_patterns:
                period_match = re.search(pattern, line, re.IGNORECASE)
                
                if period_match:
                    period_num = int(period_match.group(1))
                    rest_of_line = period_match.group(2).strip()
                    
                    # Extract time if present (multiple dash types)
                    time_match = re.search(r'(\d{1,2}):(\d{2})\s*[–\-—~]\s*(\d{1,2}):(\d{2})', rest_of_line)
                    if time_match:
                        start_time = f"{int(time_match.group(1)):02d}:{time_match.group(2)}"
                        end_time = f"{int(time_match.group(3)):02d}:{time_match.group(4)}"
                        # Remove time from rest of line to get subject
                        rest_of_line = re.sub(r'(\d{1,2}):(\d{2})\s*[–\-—~]\s*(\d{1,2}):(\d{2})', '', rest_of_line).strip()
                    else:
                        # Use default times
                        start_time, end_time = PERIOD_TIMES.get(period_num, ("08:00", "09:00"))
                    
                    # Extract subject name
                    subject = _extract_subject_name(rest_of_line)
                    
                    if subject:
                        entry = TimetableEntry(
                            day=current_day.capitalize(),
                            period_number=period_num,
                            start_time=start_time,
                            end_time=end_time,
                            subject=subject,
                            level="SSS1"
                        )
                        entries.append(entry)
                        logger.debug(f"✅ {current_day} P{period_num}: {subject} ({start_time}-{end_time})")
                        matched = True
                        break
                    else:
                        logger.warning(f"⚠️ Period found but no subject: line {line_num}: '{line}'")
            
            if not matched and line_num > 0:  # Don't warn for first line
                logger.debug(f"🔍 Line {line_num} didn't match period pattern: '{line}'")
    
    logger.info(f"📊 Structured parsing extracted {len(entries)} entries")
    return entries


def _extract_subject_name(text: str) -> str:
    """
    Extract clean subject name from text.
    Better handling of (Revision), (Practical), and combined subjects.
    """
    if not text or len(text) < 2:
        return None
    
    original_text = text
    
    # Remove time patterns if any remain
    text = re.sub(r'(\d{1,2}):(\d{2})\s*[–\-—~]\s*(\d{1,2}):(\d{2})', '', text).strip()
    
    # Handle parentheses - keep base subject
    # "Economics (Revision)" → "Economics"
    if '(' in text:
        text = text.split('(')[0].strip()
    
    # Remove anything after dash (but handle "CRS / IRS" specially)
    if ' / ' not in text and '/' not in text:
        text = re.sub(r'\s*[–—-]\s*.*$', '', text).strip()
    
    # Clean OCR artifacts
    text = _clean_ocr_text(text)
    
    # Remove extra whitespace
    text = ' '.join(text.split())
    
    # Skip if too short or too long
    if len(text) < 2 or len(text) > 60:
        logger.debug(f"❌ Subject text invalid length: '{original_text}' → '{text}'")
        return None
    
    # Check if this is a non-academic period (skip it)
    text_upper = text.upper()
    for non_academic in NON_ACADEMIC_PERIODS:
        if non_academic.upper() in text_upper:
            logger.debug(f"⏭️ Skipping non-academic period: '{original_text}'")
            return None
    
    # Try to match against known subjects
    matched = _match_subject_from_text(text)
    if matched:
        logger.debug(f"✅ Matched '{original_text}' → '{matched}'")
        return matched
    
    # If no match but text looks reasonable (has letters), return it
    if re.search(r'[a-zA-Z]{3,}', text):
        logger.debug(f"⚠️ No exact match, using: '{text}' (from '{original_text}')")
        return text.title()
    
    logger.debug(f"❌ Could not extract subject from: '{original_text}'")
    return None


def _clean_ocr_text(text: str) -> str:
    """
    Clean OCR artifacts like extra spaces within words.
    """
    if not text:
        return text
    
    subject_patterns = {
        r'mathema\s*tics': 'Mathematics',
        r'econom\s*ics': 'Economics',
        r'governm\s*ent': 'Government',
        r'commerc\s*e': 'Commerce',
        r'accountin\s*g': 'Accounting',
        r'financi\s*al': 'Financial',
        r'literatur\s*e': 'Literature',
        r'biolog\s*y': 'Biology',
        r'chemistr\s*y': 'Chemistry',
        r'physic\s*s': 'Physics',
        r'geograph\s*y': 'Geography',
        r'histor\s*y': 'History',
        r'book\s*keeping': 'Book Keeping',
        r'bookkeeping': 'Book Keeping',
    }
    
    cleaned = text
    for pattern, replacement in subject_patterns.items():
        cleaned = re.sub(pattern, replacement, cleaned, flags=re.IGNORECASE)
    
    return cleaned


def _match_subject_from_text(text: str) -> str:
    """Match text against known subjects."""
    if not text or len(text) < 3:
        return None
    
    text_clean = _clean_ocr_text(text)
    text_lower = text_clean.lower().strip()
    
    # Sort by length (longest first) to match "English Language" before "English"
    sorted_subjects = sorted(KNOWN_SUBJECTS, key=len, reverse=True)
    
    # Try exact match
    for subject in sorted_subjects:
        if subject.lower() == text_lower:
            return subject
    
    # Try partial match with word boundaries
    for subject in sorted_subjects:
        pattern = r'\b' + re.escape(subject.lower()) + r'\b'
        if re.search(pattern, text_lower):
            return subject
    
    # Try if text starts with subject
    for subject in sorted_subjects:
        if text_lower.startswith(subject.lower()):
            return subject
    
    return None


def _fallback_extraction(text: str) -> List[TimetableEntry]:
    """
    Fallback: Extract unique subjects and distribute across week.
    Only used if structured parsing fails.
    """
    logger.info("🔄 Using fallback extraction...")
    
    # Log a sample of the text to debug
    logger.debug(f"📝 Fallback text sample (first 1000 chars): {text[:1000]}")
    
    subjects = _match_known_subjects(text)
    
    if not subjects:
        logger.warning("⚠️ No known subjects matched, trying delimiter extraction...")
        subjects = _extract_from_delimited_blocks(text)
    
    if not subjects:
        logger.error("❌ No subjects found in fallback extraction")
        logger.error(f"📝 Full text: {text[:2000]}...")  # Log more text for debugging
        return []
    
    logger.info(f"✅ Found {len(subjects)} unique subjects: {subjects}")
    
    # Distribute subjects across days/periods
    entries = []
    
    subject_list = sorted(list(subjects))
    
    for idx, subject in enumerate(subject_list):
        day_idx = idx % 5
        period = (idx // 5) + 1
        
        if period > 8:  # Max 8 periods per day
            break
        
        day = DAYS_OF_WEEK[day_idx].capitalize()
        start_time, end_time = PERIOD_TIMES.get(period, ("08:00", "09:00"))
        
        entry = TimetableEntry(
            day=day,
            period_number=period,
            start_time=start_time,
            end_time=end_time,
            subject=subject,
            level="SSS1"
        )
        entries.append(entry)
        logger.debug(f"✅ Fallback: {day} P{period} = {subject}")
    
    logger.info(f"✅ Fallback created {len(entries)} entries")
    return entries


def _match_known_subjects(text: str) -> Set[str]:
    """Match complete subject names from known list."""
    subjects = set()
    
    text_clean = _clean_ocr_text(text)
    text_lower = text_clean.lower()
    
    # Sort by length (longest first) to avoid partial matches
    sorted_subjects = sorted(KNOWN_SUBJECTS, key=len, reverse=True)
    
    for subject in sorted_subjects:
        pattern = r'\b' + re.escape(subject.lower()) + r'\b'
        
        if re.search(pattern, text_lower):
            subjects.add(subject)
            logger.debug(f"✅ Matched known subject: {subject}")
    
    logger.info(f"📊 Known subjects matched: {len(subjects)}")
    return subjects


def _extract_from_delimited_blocks(text: str) -> Set[str]:
    """Extract subjects from delimited text blocks."""
    subjects = set()
    
    text_clean = _clean_ocr_text(text)
    blocks = re.split(r'[\n\t,;|]', text_clean)  # Split by common delimiters
    
    logger.debug(f"📊 Found {len(blocks)} text blocks to analyze")
    
    for block in blocks:
        block = block.strip()
        
        # Skip empty, too short, or too long blocks
        if not block or len(block) < 3 or len(block) > 60:
            continue
        
        # Remove numbers and common prefixes
        block = re.sub(r'^\d+[\.)\s]*', '', block).strip()  # Remove "1.", "1)", "1 "
        block = re.sub(r'(?:st|nd|rd|th)\s+period\s*:?\s*', '', block, flags=re.IGNORECASE).strip()
        
        if len(block) < 3:
            continue
        
        matched = _match_subject_from_text(block)
        if matched:
            subjects.add(matched)
            logger.debug(f"✅ Block matched: '{block}' → '{matched}'")
    
    logger.info(f"📊 Delimiter extraction found: {len(subjects)} subjects")
    return subjects


def _extract_from_excel(file_path: str) -> List[TimetableEntry]:
    """Extract timetable from Excel file."""
    logger.info("📊 Parsing Excel file...")
    entries = []
    
    wb = load_workbook(filename=file_path)
    
    for sheet in wb.worksheets:
        logger.info(f"📄 Processing sheet: {sheet.title}")
        
        # Try to parse as structured timetable
        sheet_entries = _parse_excel_sheet(sheet)
        entries.extend(sheet_entries)
    
    wb.close()
    
    return entries


def _parse_excel_sheet(sheet) -> List[TimetableEntry]:
    """Parse Excel sheet as timetable."""
    entries = []
    current_day = None
    
    for row in sheet.iter_rows(values_only=True):
        if not row or not any(row):
            continue
        
        # Check for day header
        first_cell = str(row[0]).strip().upper() if row[0] else ""
        for day in DAYS_OF_WEEK:
            if day in first_cell:
                current_day = day
                break
        
        # Check for period entries
        if current_day:
            for cell in row:
                if not cell:
                    continue
                
                cell_text = str(cell).strip()
                period_match = re.search(r'(\d+)(?:st|nd|rd|th)?\s+Period', cell_text, re.IGNORECASE)
                
                if period_match:
                    period_num = int(period_match.group(1))
                    subject = _extract_subject_name(cell_text)
                    
                    if subject:
                        start_time, end_time = PERIOD_TIMES.get(period_num, ("08:00", "09:00"))
                        
                        entry = TimetableEntry(
                            day=current_day.capitalize(),
                            period_number=period_num,
                            start_time=start_time,
                            end_time=end_time,
                            subject=subject,
                            level="SSS1"
                        )
                        entries.append(entry)
    
    return entries


# ============================================================
# IMAGE ENHANCEMENT FUNCTIONS (ADD THESE)
# ============================================================

def _enhance_with_clahe(image_path: str) -> str:
    """
    Apply CLAHE (Contrast Limited Adaptive Histogram Equalization) enhancement.
    Returns path to enhanced image.
    """
    img = cv2.imread(image_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Apply CLAHE
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    
    # Save enhanced image
    enhanced_path = image_path.replace('.jpg', '_clahe.jpg').replace('.jpeg', '_clahe.jpg').replace('.png', '_clahe.png')
    cv2.imwrite(enhanced_path, enhanced)
    
    return enhanced_path


def _enhance_with_binary_threshold(image_path: str) -> str:
    """Apply simple binary threshold (faster than adaptive)"""
    img = cv2.imread(image_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Simple Otsu threshold (fast)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Save
    enhanced_path = image_path.replace('.jpg', '_binary.jpg').replace('.jpeg', '_binary.jpg').replace('.png', '_binary.png')
    cv2.imwrite(enhanced_path, thresh)
    
    return enhanced_path


def _extract_from_image(file_path: str) -> List[TimetableEntry]:
    """
    Extract timetable from image using OCR with multiple strategies.
    FIXED: Adds timeouts to prevent hanging, uses best text from all strategies.
    """
    logger.info("🖼️ Processing image for extraction...")
    
    all_texts = []  # Store all extracted texts
    best_text = ""
    best_entries = []
    
    try:
        # Strategy 1: Try original image first (no enhancement) - 10 second timeout
        logger.info("📸 Strategy 1: Trying original image...")
        try:
            with time_limit(10):
                text = _try_ocr_extraction(file_path, "original")
                all_texts.append(("original", text))
                
                if len(text.strip()) > 50:
                    logger.info(f"✅ Original: {len(text)} chars")
                    entries = _parse_structured_timetable(text)
                    if entries:
                        return entries
                    if len(text) > len(best_text):
                        best_text = text
                        best_entries = entries
                else:
                    logger.warning(f"⚠️ Original: Only {len(text)} chars")
        except TimeoutException:
            logger.warning("⚠️ Strategy 1 timed out after 10s")
        except Exception as e:
            logger.warning(f"⚠️ Strategy 1 failed: {e}")
        
        # Strategy 2: Try with CLAHE - 10 second timeout
        logger.info("📸 Strategy 2: Trying CLAHE enhancement...")
        enhanced_clahe = None
        try:
            with time_limit(10):
                enhanced_clahe = _enhance_with_clahe(file_path)
                text = _try_ocr_extraction(enhanced_clahe, "CLAHE")
                all_texts.append(("CLAHE", text))
                
                if len(text.strip()) > 50:
                    logger.info(f"✅ CLAHE: {len(text)} chars")
                    entries = _parse_structured_timetable(text)
                    if entries:
                        if enhanced_clahe and os.path.exists(enhanced_clahe):
                            os.remove(enhanced_clahe)
                        return entries
                    if len(text) > len(best_text):
                        best_text = text
                        best_entries = entries
                else:
                    logger.warning(f"⚠️ CLAHE: Only {len(text)} chars")
        except TimeoutException:
            logger.warning("⚠️ Strategy 2 timed out after 10s")
        except Exception as e:
            logger.warning(f"⚠️ Strategy 2 failed: {e}")
        finally:
            if enhanced_clahe and os.path.exists(enhanced_clahe):
                try:
                    os.remove(enhanced_clahe)
                except:
                    pass
        
        # Strategy 3: SKIP adaptive threshold (it's hanging) and try binary threshold instead
        logger.info("📸 Strategy 3: Trying simple binary threshold...")
        enhanced_binary = None
        try:
            with time_limit(10):
                enhanced_binary = _enhance_with_binary_threshold(file_path)
                text = _try_ocr_extraction(enhanced_binary, "binary")
                all_texts.append(("binary", text))
                
                if len(text.strip()) > 50:
                    logger.info(f"✅ Binary: {len(text)} chars")
                    entries = _parse_structured_timetable(text)
                    if entries:
                        if enhanced_binary and os.path.exists(enhanced_binary):
                            os.remove(enhanced_binary)
                        return entries
                    if len(text) > len(best_text):
                        best_text = text
                        best_entries = entries
                else:
                    logger.warning(f"⚠️ Binary: Only {len(text)} chars")
        except TimeoutException:
            logger.warning("⚠️ Strategy 3 timed out after 10s")
        except Exception as e:
            logger.warning(f"⚠️ Strategy 3 failed: {e}")
        finally:
            if enhanced_binary and os.path.exists(enhanced_binary):
                try:
                    os.remove(enhanced_binary)
                except:
                    pass
        
        # Strategy 4: Skip denoising (slow), go straight to fallback
        logger.info("📸 All quick strategies complete")
        
        # Use the best text we got
        if best_text and len(best_text.strip()) > 20:
            logger.info(f"📝 Using best text: {len(best_text)} chars")
            
            # Try structured parsing one more time
            entries = _parse_structured_timetable(best_text)
            if entries:
                return entries
            
            # Use fallback extraction
            logger.info("🔄 Using fallback extraction on best text...")
            return _fallback_extraction(best_text)
        
        # Log all texts for debugging
        logger.error("❌ All OCR strategies failed")
        for method, text in all_texts:
            logger.debug(f"  {method}: {len(text)} chars - {text[:200]}")
        
        return []
        
    except Exception as e:
        logger.error(f"❌ Image extraction failed: {e}", exc_info=True)
        return []


def _try_ocr_extraction(image_path: str, method: str) -> str:
    """
    Try OCR extraction with multiple Tesseract configurations.
    OPTIMIZED: Only tries the best configs, faster execution.
    """
    results = []
    
    # Only try the 2 most effective configs (faster)
    
    # Config 1: PSM 6 (uniform block of text) - best for most timetables
    try:
        image = Image.open(image_path)
        text = pytesseract.image_to_string(image, config='--psm 6')
        results.append(("psm6", text))
        logger.debug(f"  {method} + PSM 6: {len(text)} chars")
    except Exception as e:
        logger.debug(f"  {method} + PSM 6 failed: {e}")
    
    # Config 2: PSM 11 (sparse text) - good for tables
    try:
        image = Image.open(image_path)
        text = pytesseract.image_to_string(image, config='--psm 11')
        results.append(("psm11", text))
        logger.debug(f"  {method} + PSM 11: {len(text)} chars")
    except Exception as e:
        logger.debug(f"  {method} + PSM 11 failed: {e}")
    
    # Return the longest result
    if results:
        best_config, best_text = max(results, key=lambda x: len(x[1]))
        logger.info(f"  Best config for {method}: {best_config} ({len(best_text)} chars)")
        return best_text
    
    return ""

def _extract_from_text_document(file_path: str) -> List[TimetableEntry]:
    """Extract from text/doc files."""
    logger.info("📄 Extracting from text document...")
    
    try:
        # For .txt files
        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        else:
            # For .doc/.docx, use python-docx or similar
            logger.warning("⚠️ .doc/.docx parsing not fully implemented")
            return []
        
        return _parse_structured_timetable(text)
        
    except Exception as e:
        logger.error(f"❌ Text extraction failed: {e}")
        return []

def _assess_image_quality(image_path: str) -> Dict[str, any]:
    """
    Assess image quality for OCR suitability.
    Returns quality score and issues found.
    """
    logger.info(f"🔍 Assessing image quality: {image_path}")
    
    try:
        # Read image
        img = cv2.imread(image_path)
        if img is None:
            logger.error(f"❌ Could not read image: {image_path}")
            return {"score": 0, "issues": ["Could not read image file"]}
        
        # Convert to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # Calculate metrics
        issues = []
        score = 100
        
        # 1. Check brightness
        mean_brightness = np.mean(gray)
        if mean_brightness < 50:
            issues.append("Image too dark")
            score -= 30
        elif mean_brightness > 200:
            issues.append("Image too bright (overexposed)")
            score -= 20
        
        logger.debug(f"   Brightness: {mean_brightness:.1f}")
        
        # 2. Check contrast (standard deviation)
        std_dev = np.std(gray)
        if std_dev < 20:
            issues.append("Low contrast - text may not be readable")
            score -= 25
        elif std_dev > 100:
            logger.debug(f"   Good contrast (std: {std_dev:.1f})")
        
        logger.debug(f"   Contrast (std dev): {std_dev:.1f}")
        
        # 3. Check sharpness (Laplacian variance)
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
        if laplacian_var < 100:
            issues.append("Image is blurry")
            score -= 30
        elif laplacian_var < 200:
            logger.debug(f"   Slight blur detected")
            score -= 10
        
        logger.debug(f"   Sharpness (Laplacian): {laplacian_var:.1f}")
        
        # 4. Check for extreme shadows/highlights
        hist = cv2.calcHist([gray], [0], None, [256], [0, 256])
        hist_normalized = hist.ravel() / hist.max()
        
        # If too many pixels at extremes (0 or 255)
        dark_pixels = (hist_normalized[0:20].sum() / len(hist_normalized)) * 100
        bright_pixels = (hist_normalized[235:256].sum() / len(hist_normalized)) * 100
        
        if dark_pixels > 30:
            issues.append("Too many dark areas (shadows)")
            score -= 15
        
        if bright_pixels > 30:
            issues.append("Too many bright areas (highlights)")
            score -= 15
        
        logger.debug(f"   Dark pixels: {dark_pixels:.1f}%, Bright pixels: {bright_pixels:.1f}%")
        
        # 5. Enhance image for better OCR
        # Normalize
        enhanced = cv2.normalize(gray, None, alpha=0, beta=255, norm_type=cv2.NORM_MINMAX)
        
        # Increase contrast using CLAHE (Contrast Limited Adaptive Histogram Equalization)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(enhanced)
        
        logger.info(f"✅ Image quality: {max(0, score)}% - {', '.join(issues) if issues else 'Good quality'}")
        
        return {
            "score": max(0, score),
            "issues": issues,
            "is_acceptable": score >= 50,
            "original": gray,
            "enhanced": enhanced
        }
        
    except Exception as e:
        logger.error(f"❌ Quality assessment failed: {e}")
        return {
            "score": 0,
            "issues": [f"Quality check error: {str(e)}"],
            "is_acceptable": False
        }


def _enhance_image_for_ocr(image_path: str) -> str:
    """
    Enhance image for better OCR results.
    Returns path to enhanced image or original if enhancement fails.
    """
    logger.info(f"🖼️ Enhancing image for OCR: {image_path}")
    
    try:
        # Read image
        img = cv2.imread(image_path)
        if img is None:
            return image_path
        
        # Convert to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # Normalize
        normalized = cv2.normalize(gray, None, alpha=0, beta=255, norm_type=cv2.NORM_MINMAX)
        
        # Apply CLAHE
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(normalized)
        
        # Thresholding for better text visibility
        _, threshold = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Save enhanced image
        enhanced_path = image_path.replace('.jpg', '_enhanced.jpg').replace('.jpeg', '_enhanced.jpg').replace('.png', '_enhanced.png')
        cv2.imwrite(enhanced_path, threshold)
        
        logger.info(f"✅ Enhanced image saved: {enhanced_path}")
        return enhanced_path
        
    except Exception as e:
        logger.error(f"❌ Image enhancement failed: {e}")
        return image_path